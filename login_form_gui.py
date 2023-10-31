import tkinter as tk
import openpyxl
from openpyxl import load_workbook
import os
from tkinter import messagebox
def F1():
    global workbook,sheet,filepath
    workbook=openpyxl.Workbook()
    sheet=workbook["Sheet"]
    filepath="C:\\Users\\Public\\Documents\\python\\Login_form_gui.xlsx"
    if len(Sv1.get())==0: messagebox.showwarning(title="Error",message="Username is Required!")
    elif len(Sv2.get())==0: messagebox.showwarning(title="Error",message="Password is Required!")
    elif not os.path.exists(filepath):
            headings=["Username","Password"]
            sheet.append(headings)
            workbook.save(filepath)
    else:
        workbook=load_workbook("Login_form_gui.xlsx")
        # sheet=workbook['Sheet']
        for i in range(len(sheet["A"])):
            if sheet["A"][i].value==Sv1.get():
                if sheet["B"][i].value!=Sv2.get():
                    messagebox.showwarning(title="Warning",message="Invaild Password\nTry Again.")
                    break
                else:
                    messagebox.showinfo(title="Login Successful",message=f"Congratulaions {Sv1.get()}")
                    break
                    
            elif i==len(sheet["A"])-1 and sheet["A"][i].value!=Sv1.get():
                messagebox.showwarning(title="Warning",message="User not Exist!")
                e3.configure(text="Sign-up",command=F2)
                break
def F2():
    workbook=load_workbook("Login_form_gui.xlsx")
    if len(Sv1.get())==0: messagebox.showwarning(title="Error",message="Username is Required!")
    elif len(Sv2.get())==0: messagebox.showwarning(title="Error",message="Password is Required!")      
    else:
        for i in range(len(sheet["A"])):
            if sheet["A"][i].value==Sv1.get():
                messagebox.showwarning(title="Error",message="User already Exists!")
            else:
                sheet.append([Sv1.get(),Sv2.get()])
                workbook.save(filepath)
                e3.configure(text=" Login ",command=F1)
                messagebox.showinfo(title="Sign-up Successful",message="Successfully Registered.")
def F3():
    pass
root=tk.Tk()
cc1="#FF1694"   #Hot pink
cc2="#28231D"   #Charcoal
Sv1=tk.StringVar()
Sv2=tk.StringVar()
root.geometry("500x500")
root.title("Login-Form")
root.configure(bg=cc2)
l1=tk.Label(root,text="Login",bg=cc2,fg=cc1,font="Courier 30 bold")
l1.pack(side=tk.TOP,pady=50)
f1=tk.Frame(root,bg=cc2)
f1.pack(side=tk.TOP,pady=10)
l2=tk.Label(f1,text="Username: ",bg=cc2,fg=cc1,font="Courier 15 bold")
l2.grid(row=0,column=0)
l3=tk.Label(f1,text="Password: ",bg=cc2,fg=cc1,font="Courier 15 bold")
l3.grid(row=1,column=0)
e1=tk.Entry(f1,textvariable=Sv1,bg=cc2,fg=cc1,font="Courier 15 bold")
e1.grid(row=0,column=1)
e2=tk.Entry(f1,textvariable=Sv2,bg=cc2,fg=cc1,font="Courier 15 bold",show="*")
e2.grid(row=1,column=1)
for i in f1.winfo_children():
     i.grid_configure(padx=10,pady=5)
e3=tk.Button(f1,text=" Login ",bg=cc2,fg=cc1,font="Courier 14 bold",activebackground=cc2,activeforeground=cc1,command=F1)
e3.grid(row=2,column=1,padx=10,pady=10)
root.mainloop()
